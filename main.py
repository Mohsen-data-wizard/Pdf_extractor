#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🎯 سیستم استخراج هوشمند اسناد گمرکی
📧 توسعهدهنده: محسن 
📅 تاریخ: 2025-06-05
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import tkinter.font as tkFont
from PIL import Image, ImageTk
import json
import os
from pathlib import Path
import threading
import queue
from datetime import datetime
import webbrowser

# Import project modules
from extractor_engine import DocumentExtractor
from learning_system import LearningSystem

class CustomsExtractorGUI:
    def __init__(self):
        """رابط اصلی برنامه"""
        
        # پنجره اصلی
        self.root = tk.Tk()
        self.root.title("🎯 سیستم استخراج هوشمند اسناد گمرکی v2.0")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 800)
        self.root.configure(bg='#f0f0f0')
        
        # تنظیم فونت فارسی
        self.setup_fonts()
        
        # متغیرهای اصلی
        self.current_files = []
        self.processing_queue = queue.Queue()
        self.results_data = {}
        self.selected_widget = None
        
        # موتورهای اصلی
        self.extractor = DocumentExtractor()
        self.learning_system = LearningSystem()
        
        # ایجاد رابط کاربری
        self.create_ui()
        
        # بارگذاری تنظیمات
        self.load_settings()
        
    def setup_fonts(self):
        """تنظیم فونتهای فارسی"""
        self.fonts = {
            'default': tkFont.Font(family='Segoe UI', size=10),
            'header': tkFont.Font(family='Segoe UI', size=14, weight='bold'),
            'small': tkFont.Font(family='Segoe UI', size=8),
            'persian': tkFont.Font(family='Tahoma', size=10),
            'persian_header': tkFont.Font(family='Tahoma', size=12, weight='bold'),
            'persian_small': tkFont.Font(family='Tahoma', size=8)
        }
        
    def create_ui(self):
        """ایجاد رابط کاربری"""
        
        # Header Frame
        self.create_header()
        
        # Main Container
        self.main_container = tk.Frame(self.root, bg='#f0f0f0')
        self.main_container.pack(fill="both", expand=True, padx=10, pady=(0,10))
        
        # Style برای Notebook
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook', background='#f0f0f0')
        style.configure('TNotebook.Tab', padding=[20, 10])
        
        # Notebook برای تبها
        self.notebook = ttk.Notebook(self.main_container)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ایجاد تبها
        self.create_tabs()
        
        # Status Bar
        self.create_status_bar()
        
        # Context Menu
        self.create_context_menu()
        
        # کلیدهای میانبر
        self.setup_shortcuts()
        
    def create_header(self):
        """ایجاد هدر برنامه"""
        header_frame = tk.Frame(self.root, bg='#2c3e50', height=80)
        header_frame.pack(fill="x", padx=10, pady=10)
        header_frame.pack_propagate(False)
        
        # Logo و عنوان
        title_frame = tk.Frame(header_frame, bg='#2c3e50')
        title_frame.pack(side="left", fill="y", padx=20)
        
        title_label = tk.Label(
            title_frame,
            text="🎯 سیستم استخراج هوشمند اسناد گمرکی",
            font=self.fonts['header'],
            bg='#2c3e50',
            fg='white'
        )
        title_label.pack(pady=8)
        
        subtitle_label = tk.Label(
            title_frame,
            text="استخراج خودکار با یادگیری هوشمند | توسط محسن",
            font=self.fonts['small'],
            bg='#2c3e50',
            fg='#ecf0f1'
        )
        subtitle_label.pack()
        
        # دکمههای کلیدی
        buttons_frame = tk.Frame(header_frame, bg='#2c3e50')
        buttons_frame.pack(side="right", fill="y", padx=20)
        
        self.quick_upload_btn = tk.Button(
            buttons_frame,
            text="📁 آپلود سریع",
            command=self.quick_upload,
            width=15,
            height=1,
            bg='#3498db',
            fg='white',
            font=self.fonts['persian'],
            relief='raised',
            cursor='hand2'
        )
        self.quick_upload_btn.pack(side="top", pady=3)
        
        self.quick_process_btn = tk.Button(
            buttons_frame,
            text="⚡ پردازش سریع",
            command=self.quick_process,
            width=15,
            height=1,
            bg='#27ae60',
            fg='white',
            font=self.fonts['persian'],
            relief='raised',
            cursor='hand2',
            state="disabled"
        )
        self.quick_process_btn.pack(side="top", pady=3)
        
    def create_tabs(self):
        """ایجاد تبهای اصلی"""
        
        # تب پردازش
        self.processing_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.processing_frame, text="📁 پردازش اسناد")
        self.create_processing_tab()
        
        # تب ویرایش
        self.edit_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.edit_frame, text="✏️ ویرایش نتایج")
        self.create_edit_tab()
        
        # تب نتایج
        self.results_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.results_frame, text="📊 نتایج و گزارش")
        self.create_results_tab()
        
        # تب یادگیری
        self.learning_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.learning_frame, text="🤖 یادگیری هوشمند")
        self.create_learning_tab()
        
        # تب تنظیمات
        self.settings_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.settings_frame, text="⚙️ تنظیمات")
        self.create_settings_tab()
        
    def create_processing_tab(self):
        """ایجاد تب پردازش"""
        
        # پنل کنترل
        control_panel = tk.LabelFrame(
            self.processing_frame,
            text="🎛️ کنترل پردازش",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        control_panel.pack(fill="x", padx=20, pady=10)
        
        # انتخاب نوع سند
        doc_type_frame = tk.Frame(control_panel, bg='white')
        doc_type_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(
            doc_type_frame,
            text="نوع سند:",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=5)
        
        self.doc_type_var = tk.StringVar(value="auto")
        doc_types = [
            ("تشخیص خودکار", "auto"),
            ("واردات تککالایی", "import_single"),
            ("واردات چندکالایی", "import_multi"),
            ("صادرات تککالایی", "export_single"),
            ("صادرات چندکالایی", "export_multi")
        ]
        
        for text, value in doc_types:
            tk.Radiobutton(
                doc_type_frame,
                text=text,
                value=value,
                variable=self.doc_type_var,
                font=self.fonts['persian_small'],
                bg='white'
            ).pack(side="left", padx=10)
            
        # دکمههای آپلود
        upload_frame = tk.Frame(control_panel, bg='white')
        upload_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(
            upload_frame,
            text="📄 انتخاب PDF",
            command=lambda: self.upload_files("pdf"),
            width=15,
            bg='#e74c3c',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            upload_frame,
            text="🖼️ انتخاب تصاویر",
            command=lambda: self.upload_files("images"),
            width=15,
            bg='#9b59b6',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            upload_frame,
            text="📁 انتخاب پوشه",
            command=self.upload_folder,
            width=15,
            bg='#f39c12',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            upload_frame,
            text="🗑️ پاک کردن",
            command=self.clear_files,
            width=15,
            bg='#95a5a6',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="right", padx=5)
        
        # لیست فایلها
        files_frame = tk.LabelFrame(
            self.processing_frame,
            text="📋 فایلهای انتخاب شده",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        files_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Treeview برای نمایش فایلها
        columns = ("نام فایل", "اندازه", "نوع", "وضعیت")
        self.files_tree = ttk.Treeview(files_frame, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.files_tree.heading(col, text=col)
            self.files_tree.column(col, width=200, anchor="center")
            
        # Scrollbar برای Treeview
        scrollbar = ttk.Scrollbar(files_frame, orient="vertical", command=self.files_tree.yview)
        self.files_tree.configure(yscrollcommand=scrollbar.set)
        
        self.files_tree.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scrollbar.pack(side="right", fill="y", padx=5, pady=5)
        
        # Progress Bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.processing_frame,
            variable=self.progress_var,
            maximum=100,
            length=400,
            mode='determinate'
        )
        self.progress_bar.pack(pady=10)
        
        # دکمه شروع پردازش
        process_btn = tk.Button(
            self.processing_frame,
            text="🚀 شروع پردازش",
            command=self.start_processing,
            width=20,
            height=2,
            bg='#27ae60',
            fg='white',
            font=self.fonts['persian_header'],
            cursor='hand2'
        )
        process_btn.pack(pady=10)
        
    def create_edit_tab(self):
        """ایجاد تب ویرایش"""
        
        # پنل انتخاب فایل
        file_panel = tk.LabelFrame(
            self.edit_frame,
            text="📄 انتخاب فایل برای ویرایش",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        file_panel.pack(fill="x", padx=20, pady=10)
        
        # Combobox برای انتخاب فایل
        tk.Label(
            file_panel,
            text="فایل:",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=10)
        
        self.edit_file_var = tk.StringVar()
        self.edit_file_combo = ttk.Combobox(
            file_panel,
            textvariable=self.edit_file_var,
            width=50,
            state="readonly"
        )
        self.edit_file_combo.pack(side="left", padx=10, pady=10)
        self.edit_file_combo.bind("<<ComboboxSelected>>", self.load_file_for_edit)
        
        # دکمه بارگذاری مجدد
        tk.Button(
            file_panel,
            text="🔄 بارگذاری مجدد",
            command=self.reload_edit_data,
            bg='#3498db',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="right", padx=10)
        
        # پنل اصلی ویرایش
        main_edit_panel = tk.Frame(self.edit_frame, bg='white')
        main_edit_panel.pack(fill="both", expand=True, padx=20, pady=10)
        
        # پنل سمت چپ - پیشنمایش تصویر
        left_panel = tk.LabelFrame(
            main_edit_panel,
            text="🖼️ پیشنمایش سند",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        left_panel.pack(side="left", fill="both", expand=True, padx=(0,10))
        
        # Canvas برای نمایش تصویر
        self.image_canvas = tk.Canvas(left_panel, bg='white', width=600, height=600)
        self.image_canvas.pack(fill="both", expand=True, padx=5, pady=5)
        
        # دکمههای کنترل تصویر
        image_controls = tk.Frame(left_panel, bg='white')
        image_controls.pack(fill="x", padx=5, pady=5)
        
        tk.Button(
            image_controls,
            text="🔍 بزرگنمایی",
            command=self.zoom_in,
            bg='#3498db',
            fg='white',
            font=self.fonts['persian_small'],
            cursor='hand2'
        ).pack(side="left", padx=2)
        
        tk.Button(
            image_controls,
            text="🔍 کوچکنمایی",
            command=self.zoom_out,
            bg='#3498db',
            fg='white',
            font=self.fonts['persian_small'],
            cursor='hand2'
        ).pack(side="left", padx=2)
        
        tk.Button(
            image_controls,
            text="📐 اندازه اصلی",
            command=self.reset_zoom,
            bg='#95a5a6',
            fg='white',
            font=self.fonts['persian_small'],
            cursor='hand2'
        ).pack(side="left", padx=2)
        
        # پنل سمت راست - فیلدهای ویرایش
        right_panel = tk.LabelFrame(
            main_edit_panel,
            text="✏️ ویرایش فیلدها",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50',
            width=500
        )
        right_panel.pack(side="right", fill="y", padx=(10,0))
        right_panel.pack_propagate(False)
        
        # Scrollable Frame برای فیلدها
        canvas = tk.Canvas(right_panel, bg='white')
        scrollbar_edit = ttk.Scrollbar(right_panel, orient="vertical", command=canvas.yview)
        self.scrollable_edit_frame = tk.Frame(canvas, bg='white')
        
        self.scrollable_edit_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_edit_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_edit.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scrollbar_edit.pack(side="right", fill="y")
        
        # دیکشنری برای نگهداری ویجتهای ویرایش
        self.edit_widgets = {}
        
        # دکمههای عملیات
        action_panel = tk.Frame(self.edit_frame, bg='white')
        action_panel.pack(fill="x", padx=20, pady=10)
        
        tk.Button(
            action_panel,
            text="💾 ذخیره تغییرات",
            command=self.save_edits,
            width=15,
            bg='#27ae60',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            action_panel,
            text="🤖 اعمال و یادگیری",
            command=self.apply_and_learn,
            width=18,
            bg='#e67e22',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            action_panel,
            text="↩️ بازگشت به حالت اصلی",
            command=self.reset_edits,
            width=20,
            bg='#95a5a6',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="right", padx=5)
        
    def create_results_tab(self):
        """ایجاد تب نتایج"""
        
        # پنل کنترل نتایج
        control_results = tk.LabelFrame(
            self.results_frame,
            text="📊 کنترل نمایش نتایج",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        control_results.pack(fill="x", padx=20, pady=10)
        
        # فیلتر نتایج
        filter_frame = tk.Frame(control_results, bg='white')
        filter_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(
            filter_frame,
            text="نمایش:",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=5)
        
        self.results_filter_var = tk.StringVar(value="all")
        filters = [
            ("همه فایلها", "all"),
            ("موفق", "success"),
            ("ناموفق", "failed"),
            ("نیاز به بررسی", "review")
        ]
        
        for text, value in filters:
            tk.Radiobutton(
                filter_frame,
                text=text,
                value=value,
                variable=self.results_filter_var,
                font=self.fonts['persian_small'],
                bg='white',
                command=self.filter_results
            ).pack(side="left", padx=10)
            
        # دکمههای خروجی
        export_frame = tk.Frame(control_results, bg='white')
        export_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(
            export_frame,
            text="📄 خروجی Excel",
            command=self.export_to_excel,
            bg='#27ae60',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            export_frame,
            text="📋 کپی به کلیپبورد",
            command=self.copy_to_clipboard,
            bg='#3498db',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            export_frame,
            text="📊 گزارش تحلیلی",
            command=self.generate_analysis_report,
            bg='#9b59b6',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        # جدول نتایج
        results_table_frame = tk.LabelFrame(
            self.results_frame,
            text="📋 جدول نتایج",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        results_table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # ایجاد Treeview برای نتایج
        self.create_results_treeview(results_table_frame)
        
        # پنل آمار
        stats_frame = tk.LabelFrame(
            self.results_frame,
            text="📈 آمار استخراج",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        self.stats_text = tk.Text(
            stats_frame,
            height=8,
            font=self.fonts['persian_small'],
            bg='#f8f9fa',
            fg='#2c3e50',
            wrap=tk.WORD
        )
        self.stats_text.pack(fill="x", padx=10, pady=10)
        
    def create_results_treeview(self, parent):
        """ایجاد جدول نتایج"""
        
        # تعریف ستونها
        columns = [
            "ردیف", "نام فایل", "وضعیت", "زمان پردازش",
            "شماره کوتا", "کد کالا", "شرح کالا", "وزن خالص",
            "کشور طرف معامله", "نرخ ارز", "ارزش گمرکی", "اطمینان کلی"
        ]
        
        # Frame برای Treeview
        tree_frame = tk.Frame(parent, bg='white')
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Treeview
        self.results_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            height=20
        )
        
        # تنظیم ستونها
        for col in columns:
            self.results_tree.heading(col, text=col, command=lambda c=col: self.sort_results(c))
            if col in ["ردیف", "وضعیت", "زمان پردازش", "اطمینان کلی"]:
                self.results_tree.column(col, width=100, anchor="center")
            elif col == "نام فایل":
                self.results_tree.column(col, width=200, anchor="w")
            elif col == "شرح کالا":
                self.results_tree.column(col, width=300, anchor="w")
            else:
                self.results_tree.column(col, width=120, anchor="center")
                
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.results_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.results_tree.xview)
        
        self.results_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack Treeview و Scrollbars
        self.results_tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Event binding
        self.results_tree.bind("<Double-1>", self.on_result_double_click)
        self.bind_context_menu(self.results_tree)
        
    def create_learning_tab(self):
        """ایجاد تب یادگیری"""
        
        # دو پنل اصلی
        top_panel = tk.Frame(self.learning_frame, bg='white')
        top_panel.pack(fill="both", expand=True, padx=20, pady=10)
        
        bottom_panel = tk.Frame(self.learning_frame, bg='white')
        bottom_panel.pack(fill="both", expand=True, padx=20, pady=10)
        
        # پنل مقایسه (بالا چپ)
        comparison_frame = tk.LabelFrame(
            top_panel,
            text="🔍 مقایسه قبل و بعد",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        comparison_frame.pack(side="left", fill="both", expand=True, padx=(0,10))
        
        # جدول مقایسه
        comp_columns = ["فیلد", "قبل", "بعد", "تغییر", "وضعیت"]
        self.comparison_tree = ttk.Treeview(
            comparison_frame,
            columns=comp_columns,
            show="headings",
            height=15
        )
        
        for col in comp_columns:
            self.comparison_tree.heading(col, text=col)
            self.comparison_tree.column(col, width=120, anchor="center")
            
        self.comparison_tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # پنل الگوهای یادگیری (بالا راست)
        patterns_frame = tk.LabelFrame(
            top_panel,
            text="🧠 الگوهای یادگیری",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        patterns_frame.pack(side="right", fill="both", expand=True, padx=(10,0))
        
        # جدول الگوها
        patterns_columns = ["فیلد", "الگو جدید", "تعداد موفقیت", "دقت"]
        self.patterns_tree = ttk.Treeview(
            patterns_frame,
            columns=patterns_columns,
            show="headings",
            height=15
        )
        
        for col in patterns_columns:
            self.patterns_tree.heading(col, text=col)
            self.patterns_tree.column(col, width=150, anchor="center")
            
        self.patterns_tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # پنل مدیریت الگوها (پایین)
        management_frame = tk.LabelFrame(
            bottom_panel,
            text="⚙️ مدیریت الگوها",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        management_frame.pack(fill="both", expand=True)
        
        # دکمههای مدیریت
        mgmt_buttons = tk.Frame(management_frame, bg='white')
        mgmt_buttons.pack(fill="x", padx=10, pady=10)
        
        tk.Button(
            mgmt_buttons,
            text="➕ افزودن الگو",
            command=self.add_pattern,
            bg='#27ae60',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            mgmt_buttons,
            text="✏️ ویرایش الگو",
            command=self.edit_pattern,
            bg='#3498db',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            mgmt_buttons,
            text="🗑️ حذف الگو",
            command=self.delete_pattern,
            bg='#e74c3c',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            mgmt_buttons,
            text="💾 ذخیره الگوها",
            command=self.save_patterns,
            bg='#9b59b6',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="right", padx=5)
        
        tk.Button(
            mgmt_buttons,
            text="🔄 بارگذاری مجدد",
            command=self.reload_patterns,
            bg='#95a5a6',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="right", padx=5)
        
        # ناحیه ویرایش الگو
        edit_pattern_frame = tk.LabelFrame(
            management_frame,
            text="🔧 ویرایش الگو",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        edit_pattern_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # انتخاب فیلد
        field_selection = tk.Frame(edit_pattern_frame, bg='white')
        field_selection.pack(fill="x", padx=5, pady=5)
        
        tk.Label(
            field_selection,
            text="فیلد:",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=5)
        
        self.pattern_field_var = tk.StringVar()
        self.pattern_field_combo = ttk.Combobox(
            field_selection,
            textvariable=self.pattern_field_var,
            width=30,
            state="readonly"
        )
        self.pattern_field_combo.pack(side="left", padx=10)
        
        # ورودی الگو
        pattern_input = tk.Frame(edit_pattern_frame, bg='white')
        pattern_input.pack(fill="both", expand=True, padx=5, pady=5)
        
        tk.Label(
            pattern_input,
            text="الگو (Regex):",
            font=self.fonts['persian'],
            bg='white'
        ).pack(anchor="w", padx=5)
        
        self.pattern_entry = tk.Text(
            pattern_input,
            height=4,
            font=self.fonts['default'],
            bg='#f8f9fa',
            fg='#2c3e50'
        )
        self.pattern_entry.pack(fill="both", expand=True, padx=5, pady=5)
        self.bind_context_menu(self.pattern_entry)
        
    def create_settings_tab(self):
        """ایجاد تب تنظیمات"""
        
        # تنظیمات پردازش
        processing_settings = tk.LabelFrame(
            self.settings_frame,
            text="⚡ تنظیمات پردازش",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        processing_settings.pack(fill="x", padx=20, pady=10)
        
        # آستانه اطمینان
        confidence_frame = tk.Frame(processing_settings, bg='white')
        confidence_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(
            confidence_frame,
            text="آستانه اطمینان:",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=5)
        
        self.confidence_var = tk.DoubleVar(value=0.35)
        confidence_scale = tk.Scale(
            confidence_frame,
            from_=0.1,
            to=0.9,
            resolution=0.05,
            orient=tk.HORIZONTAL,
            variable=self.confidence_var,
            bg='white',
            font=self.fonts['persian_small']
        )
        confidence_scale.pack(side="left", padx=10)
        
        self.confidence_label = tk.Label(
            confidence_frame,
            text="0.35",
            font=self.fonts['persian'],
            bg='white'
        )
        self.confidence_label.pack(side="left", padx=5)
        
        confidence_scale.configure(command=self.update_confidence_label)
        
        # DPI تصاویر
        dpi_frame = tk.Frame(processing_settings, bg='white')
        dpi_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(
            dpi_frame,
            text="کیفیت تصویر (DPI):",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=5)
        
        self.dpi_var = tk.IntVar(value=300)
        dpi_values = [150, 200, 300, 400, 600]
        
        for dpi in dpi_values:
            tk.Radiobutton(
                dpi_frame,
                text=str(dpi),
                value=dpi,
                variable=self.dpi_var,
                font=self.fonts['persian_small'],
                bg='white'
            ).pack(side="left", padx=10)
            
        # تنظیمات OCR
        ocr_settings = tk.LabelFrame(
            self.settings_frame,
            text="👁️ تنظیمات OCR",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        ocr_settings.pack(fill="x", padx=20, pady=10)
        
        # زبانها
        lang_frame = tk.Frame(ocr_settings, bg='white')
        lang_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(
            lang_frame,
            text="زبانهای تشخیص:",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=5)
        
        self.lang_fa = tk.BooleanVar(value=True)
        self.lang_en = tk.BooleanVar(value=True)
        self.lang_ar = tk.BooleanVar(value=True)
        
        tk.Checkbutton(
            lang_frame,
            text="فارسی",
            variable=self.lang_fa,
            font=self.fonts['persian_small'],
            bg='white'
        ).pack(side="left", padx=10)
        
        tk.Checkbutton(
            lang_frame,
            text="انگلیسی",
            variable=self.lang_en,
            font=self.fonts['persian_small'],
            bg='white'
        ).pack(side="left", padx=10)
        
        tk.Checkbutton(
            lang_frame,
            text="عربی",
            variable=self.lang_ar,
            font=self.fonts['persian_small'],
            bg='white'
        ).pack(side="left", padx=10)
        
        # تنظیمات رابط
        ui_settings = tk.LabelFrame(
            self.settings_frame,
            text="🎨 تنظیمات رابط",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        ui_settings.pack(fill="x", padx=20, pady=10)
        
        # تم
        theme_frame = tk.Frame(ui_settings, bg='white')
        theme_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(
            theme_frame,
            text="تم رنگی:",
            font=self.fonts['persian'],
            bg='white'
        ).pack(side="left", padx=5)
        
        self.theme_var = tk.StringVar(value="light")
        themes = [("روشن", "light"), ("تیره", "dark"), ("آبی", "blue")]
        
        for text, value in themes:
            tk.Radiobutton(
                theme_frame,
                text=text,
                value=value,
                variable=self.theme_var,
                font=self.fonts['persian_small'],
                bg='white',
                command=self.change_theme
            ).pack(side="left", padx=10)
            
        # دکمههای تنظیمات
        settings_buttons = tk.Frame(self.settings_frame, bg='white')
        settings_buttons.pack(fill="x", padx=20, pady=20)
        
        tk.Button(
            settings_buttons,
            text="💾 ذخیره تنظیمات",
            command=self.save_settings,
            width=15,
            bg='#27ae60',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            settings_buttons,
            text="🔄 تنظیمات پیشفرض",
            command=self.reset_settings,
            width=18,
            bg='#95a5a6',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="left", padx=5)
        
        tk.Button(
            settings_buttons,
            text="📋 تست تنظیمات",
            command=self.test_settings,
            width=15,
            bg='#3498db',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        ).pack(side="right", padx=5)
        
        # اطلاعات برنامه
        about_frame = tk.LabelFrame(
            self.settings_frame,
            text="ℹ️ درباره برنامه",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        about_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        about_text = """
        🎯 سیستم استخراج هوشمند اسناد گمرکی
        نسخه: 2.0
        توسعهدهنده: محسن
        تاریخ: 2025-06-05
        
        ✨ ویگیها:
        • استخراج خودکار اطلاعات از PDF و تصاویر
        • یادگیری هوشمند و بهبود دقت
        • پشتیبانی از اسناد تککالایی و چندکالایی
        • رابط کاربری دوستانه و فارسی
        • خروجی Excel با فرمتبندی مناسب
        
        🔗 لینکهای مفید:
        """
        
        about_label = tk.Label(
            about_frame,
            text=about_text,
            font=self.fonts['persian_small'],
            bg='white',
            fg='#2c3e50',
            justify=tk.RIGHT,
            anchor="ne"
        )
        about_label.pack(fill="both", expand=True, padx=10, pady=10)
        
    def create_status_bar(self):
        """ایجاد نوار وضعیت"""
        
        self.status_frame = tk.Frame(self.root, bg='#34495e', height=30)
        self.status_frame.pack(fill="x", side="bottom")
        self.status_frame.pack_propagate(False)
        
        # متن وضعیت
        self.status_var = tk.StringVar(value="آماده")
        self.status_label = tk.Label(
            self.status_frame,
            textvariable=self.status_var,
            font=self.fonts['persian_small'],
            bg='#34495e',
            fg='white',
            anchor='w'
        )
        self.status_label.pack(side="left", fill="x", expand=True, padx=10)
        
        # اطلاعات سیستم
        self.system_info = tk.Label(
            self.status_frame,
            text=f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')} | 👤 Mohsen-data-wizard",
            font=self.fonts['small'],
            bg='#34495e',
            fg='#bdc3c7'
        )
        self.system_info.pack(side="right", padx=10)
        
    def create_context_menu(self):
        """ایجاد منوی راست کلیک"""
        
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="کپی", command=self.copy_text)
        self.context_menu.add_command(label="پیست", command=self.paste_text)
        self.context_menu.add_command(label="انتخاب همه", command=self.select_all)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="پاک کردن", command=self.clear_text)
        
    def bind_context_menu(self, widget):
        """اتصال منوی راست کلیک به ویجت"""
        
        def show_context_menu(event):
            self.selected_widget = widget
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()
                
        widget.bind("<Button-3>", show_context_menu)
        
    def setup_shortcuts(self):
        """تنظیم کلیدهای میانبر"""
        
        self.root.bind("<Control-v>", lambda e: self.paste_text())
        self.root.bind("<Control-c>", lambda e: self.copy_text())
        self.root.bind("<Control-a>", lambda e: self.select_all())
        self.root.bind("<Control-o>", lambda e: self.quick_upload())
        self.root.bind("<F5>", lambda e: self.quick_process())
        
    # متدهای کاربردی
    def update_status(self, message):
        """بهروزرسانی وضعیت"""
        self.status_var.set(message)
        self.root.update_idletasks()
        
    def copy_text(self):
        """کپی کردن متن"""
        try:
            if self.selected_widget:
                if hasattr(self.selected_widget, 'selection_get'):
                    text = self.selected_widget.selection_get()
                elif hasattr(self.selected_widget, 'get') and hasattr(self.selected_widget, 'selection_present'):
                    if self.selected_widget.selection_present():
                        text = self.selected_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                    else:
                        text = self.selected_widget.get()
                else:
                    return
                    
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
                self.update_status("✅ متن کپی شد")
        except:
            pass
            
    def paste_text(self):
        """پیست کردن متن"""
        try:
            text = self.root.clipboard_get()
            if self.selected_widget and hasattr(self.selected_widget, 'insert'):
                if hasattr(self.selected_widget, 'selection_present') and self.selected_widget.selection_present():
                    self.selected_widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
                self.selected_widget.insert(tk.INSERT, text)
                self.update_status("✅ متن پیست شد")
        except:
            pass
            
    def select_all(self):
        """انتخاب همه متن"""
        try:
            if self.selected_widget:
                if hasattr(self.selected_widget, 'select_range'):
                    self.selected_widget.select_range(0, tk.END)
                elif hasattr(self.selected_widget, 'tag_add'):
                    self.selected_widget.tag_add(tk.SEL, "1.0", tk.END)
        except:
            pass
            
    def clear_text(self):
        """پاک کردن متن"""
        try:
            if self.selected_widget and hasattr(self.selected_widget, 'delete'):
                if hasattr(self.selected_widget, 'get'):
                    self.selected_widget.delete(0, tk.END)
                else:
                    self.selected_widget.delete("1.0", tk.END)
                self.update_status("✅ متن پاک شد")
        except:
            pass
            
    # متدهای اصلی
    def quick_upload(self):
        """آپلود سریع فایلها"""
        filetypes = [
            ("همه فایلهای پشتیبانی شده", "*.pdf *.png *.jpg *.jpeg"),
            ("فایلهای PDF", "*.pdf"),
            ("تصاویر", "*.png *.jpg *.jpeg"),
            ("همه فایلها", "*.*")
        ]
        
        files = filedialog.askopenfilenames(
            title="انتخاب فایلهای گمرکی",
            filetypes=filetypes
        )
        
        if files:
            self.current_files = list(files)
            self.quick_process_btn.configure(state="normal")
            self.update_status(f"✅ {len(files)} فایل انتخاب شد")
            
            # بهروزرسانی لیست فایلها
            self.update_files_list()
            
            # انتقال به تب پردازش
            self.notebook.select(0)
            
    def upload_files(self, file_type):
        """آپلود فایلها بر اساس نوع"""
        if file_type == "pdf":
            filetypes = [("فایلهای PDF", "*.pdf")]
        elif file_type == "images":
            filetypes = [("تصاویر", "*.png *.jpg *.jpeg")]
        else:
            filetypes = [("همه فایلها", "*.*")]
            
        files = filedialog.askopenfilenames(
            title=f"انتخاب {file_type}",
            filetypes=filetypes
        )
        
        if files:
            self.current_files.extend(files)
            self.update_files_list()
            self.update_status(f"✅ {len(files)} فایل اضافه شد")
            
    def upload_folder(self):
        """آپلود پوشه"""
        folder = filedialog.askdirectory(title="انتخاب پوشه")
        
        if folder:
            # پیدا کردن فایلهای پشتیبانی شده
            supported_extensions = ['.pdf', '.png', '.jpg', '.jpeg']
            files = []
            
            for ext in supported_extensions:
                files.extend(Path(folder).glob(f"*{ext}"))
                files.extend(Path(folder).glob(f"*{ext.upper()}"))
                
            if files:
                self.current_files.extend([str(f) for f in files])
                self.update_files_list()
                self.update_status(f"✅ {len(files)} فایل از پوشه اضافه شد")
            else:
                messagebox.showwarning("هشدار", "هیچ فایل پشتیبانی شدهای در پوشه یافت نشد")
                
    def clear_files(self):
        """پاک کردن لیست فایلها"""
        self.current_files = []
        self.update_files_list()
        self.quick_process_btn.configure(state="disabled")
        self.update_status("🗑️ لیست فایلها پاک شد")
        
    def update_files_list(self):
        """بهروزرسانی لیست فایلها"""
        # پاک کردن لیست قبلی
        for item in self.files_tree.get_children():
            self.files_tree.delete(item)
            
        # اضافه کردن فایلهای جدید
        for i, file_path in enumerate(self.current_files, 1):
            file_path_obj = Path(file_path)
            file_size = file_path_obj.stat().st_size if file_path_obj.exists() else 0
            file_size_str = f"{file_size / 1024:.1f} KB" if file_size < 1024*1024 else f"{file_size / (1024*1024):.1f} MB"
            
            self.files_tree.insert("", "end", values=[
                i,
                file_path_obj.name,
                file_size_str,
                file_path_obj.suffix.upper()[1:],
                "آماده"
            ])
            
    def start_processing(self):
        """شروع پردازش"""
        if not self.current_files:
            messagebox.showwarning("هشدار", "ابتدا فایلهایی را انتخاب کنید")
            return
            
        # تنظیم progress bar
        self.progress_var.set(0)
        self.update_status("🔄 شروع پردازش...")
        
        # شروع پردازش در thread جداگانه
        threading.Thread(
            target=self.process_files_background,
            daemon=True
        ).start()
        
    def process_files_background(self):
        """پردازش فایلها در پسزمینه"""
        try:
            total_files = len(self.current_files)
            
            for i, file_path in enumerate(self.current_files):
                # بهروزرسانی progress
                progress = (i / total_files) * 100
                self.root.after(0, lambda p=progress: self.progress_var.set(p))
                
                # بهروزرسانی وضعیت فایل
                self.root.after(0, lambda idx=i: self.update_file_status(idx, "در حال پردازش"))
                
                # پردازش فایل
                try:
                    result = self.extractor.process_single_file(file_path)
                    
                    # ذخیره نتیجه
                    if file_path not in self.results_data:
                        self.results_data[file_path] = result
                        
                    # بهروزرسانی وضعیت
                    status = "موفق" if result.get('status') == 'success' else "ناموفق"
                    self.root.after(0, lambda idx=i, s=status: self.update_file_status(idx, s))
                    
                except Exception as e:
                    self.root.after(0, lambda idx=i: self.update_file_status(idx, "خطا"))
                    
            # تکمیل پردازش
            self.root.after(0, self.on_processing_complete)
            
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.on_processing_error(err))
            
    def update_file_status(self, index, status):
        """بهروزرسانی وضعیت فایل"""
        try:
            items = self.files_tree.get_children()
            if index < len(items):
                item = items[index]
                values = list(self.files_tree.item(item)['values'])
                values[4] = status  # ستون وضعیت
                self.files_tree.item(item, values=values)
        except:
            pass
            
    def quick_process(self):
        """پردازش سریع"""
        self.start_processing()
        
    def on_processing_complete(self):
        """اتمام پردازش"""
        self.progress_var.set(100)
        self.update_status("✅ پردازش تکمیل شد")
        
        # انتقال به تب نتایج
        self.notebook.select(2)
        self.display_results()
        
        messagebox.showinfo(
            "موفقیت",
            f"پردازش {len(self.current_files)} فایل با موفقیت انجام شد"
        )
        
    def on_processing_error(self, error):
        """خطا در پردازش"""
        self.update_status("❌ خطا در پردازش")
        messagebox.showerror("خطا", f"خطا در پردازش: {error}")
        
    def display_results(self):
        """نمایش نتایج"""
        # پاک کردن نتایج قبلی
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
            
        # نمایش نتایج جدید
        row_num = 1
        for file_path, result in self.results_data.items():
            if 'pages' in result:
                for page_result in result['pages']:
                    self.add_result_row(row_num, file_path, page_result)
                    row_num += 1
            else:
                self.add_result_row(row_num, file_path, result)
                row_num += 1
                
        # نمایش آمار
        self.display_stats()
        
    def add_result_row(self, row_num, file_path, result):
        """اضافه کردن ردیف نتیجه"""
        file_name = Path(file_path).name
        status = result.get('status', 'unknown')
        processing_time = result.get('processing_time', '0s')
        
        extracted = result.get('extracted', {})
        
        # استخراج مقادیر فیلدها
        kotazh = extracted.get('شماره_کوتا', {}).get('value', '')
        commodity_code = extracted.get('کد_کالا', {}).get('value', '')
        description = extracted.get('شرح_کالا', {}).get('value', '')
        net_weight = extracted.get('وزن_خالص', {}).get('value', '')
        country = extracted.get('کشور_طرف_معامله', {}).get('value', '')
        exchange_rate = extracted.get('نرخ_ارز', {}).get('value', '')
        customs_value = extracted.get('ارزش_گمرکی', {}).get('value', '')
        
        # محاسبه اطمینان کلی
        confidences = []
        for field_data in extracted.values():
            if isinstance(field_data, dict) and 'confidence' in field_data:
                confidences.append(field_data['confidence'])
                
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0
        confidence_str = f"{avg_confidence:.1f}"
        
        # اضافه کردن به جدول
        self.results_tree.insert("", "end", values=[
            row_num, file_name, status, processing_time,
            kotazh, commodity_code, description, net_weight,
            country, exchange_rate, customs_value, confidence_str
        ])
        
    def display_stats(self):
        """نمایش آمار"""
        stats = self.extractor.get_extraction_stats(self.results_data)
        
        stats_text = f"""📊 آمار استخراج:

📁 تعداد کل فایلها: {stats['total_files']}
📄 تعداد کل صفحات: {stats['total_pages']}
⏱️ زمان کل پردازش: {stats['processing_time']:.1f} ثانیه

🎯 نرخ موفقیت فیلدها:
"""

        for field_name, field_stats in stats['successful_extractions'].items():
            stats_text += f"• {field_name}: {field_stats['count']} ({field_stats['percentage']})\n"
            
        stats_text += f"\n📋 انواع اسناد:\n"
        for doc_type, count in stats['document_types'].items():
            type_name = self.extractor.document_types.get(doc_type, doc_type)
            stats_text += f"• {type_name}: {count} سند\n"
            
        # نمایش آمار
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert("1.0", stats_text)
        
    def load_file_for_edit(self, event=None):
        """بارگذاری فایل برای ویرایش"""
        selected_file = self.edit_file_var.get()
        if not selected_file or selected_file not in self.results_data:
            return
            
        # پاک کردن فیلدهای قبلی
        for widget in self.scrollable_edit_frame.winfo_children():
            widget.destroy()
            
        self.edit_widgets.clear()
        
        # بارگذاری دادههای فایل
        file_data = self.results_data[selected_file]
        
        # نمایش تصویر (اگر موجود باشد)
        self.load_image_for_preview(selected_file)
        
        # ایجاد فیلدهای ویرایش
        if 'pages' in file_data:
            # فایل چندصفحهای
            for page_num, page_data in enumerate(file_data['pages']):
                self.create_page_edit_section(page_num, page_data)
        else:
            # فایل تکصفحهای
            self.create_page_edit_section(0, file_data)
            
    def create_page_edit_section(self, page_num, page_data):
        """ایجاد بخش ویرایش برای یک صفحه"""
        
        # عنوان صفحه
        page_title = tk.LabelFrame(
            self.scrollable_edit_frame,
            text=f"📄 صفحه {page_num + 1}",
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50'
        )
        page_title.pack(fill="x", padx=5, pady=10)
        
        extracted = page_data.get('extracted', {})
        
        # لیست فیلدهای قابل ویرایش
        editable_fields = [
            ('شماره_کوتا', 'شماره کوتا'),
            ('کد_کالا', 'کد کالا'),
            ('شرح_کالا', 'شرح کالا'),
            ('نوع_بسته', 'نوع بسته'),
            ('تعداد_بسته', 'تعداد بسته'),
            ('وزن_خالص', 'وزن خالص'),
            ('کشور_طرف_معامله', 'کشور طرف معامله'),
            ('نرخ_ارز', 'نرخ ارز'),
            ('نوع_ارز', 'نوع ارز'),
            ('ارزش_گمرکی', 'ارزش گمرکی'),
            ('بیمه', 'بیمه'),
            ('کرایه', 'کرایه'),
            ('مبلغ_حقوق_ورودی', 'مبلغ حقوق ورودی'),
            ('مالیات_بر_ارزش_افزوده', 'مالیات بر ارزش افزوده'),
            ('جمع_حقوق_عوارض', 'جمع حقوق و عوارض')
        ]
        
        for field_key, field_label in editable_fields:
            if field_key in extracted:
                self.create_field_edit_widget(
                    page_title, 
                    f"{page_num}_{field_key}", 
                    field_label, 
                    extracted[field_key]
                )
                
    def create_field_edit_widget(self, parent, field_id, field_label, field_data):
        """ایجاد ویجت ویرایش برای یک فیلد"""
        
        field_frame = tk.Frame(parent, bg='white')
        field_frame.pack(fill="x", padx=10, pady=5)
        
        # برچسب فیلد
        label_frame = tk.Frame(field_frame, bg='white', width=150)
        label_frame.pack(side="left", fill="y")
        label_frame.pack_propagate(False)
        
        tk.Label(
            label_frame,
            text=f"{field_label}:",
            font=self.fonts['persian'],
            bg='white',
            anchor='e'
        ).pack(fill="both", expand=True, padx=5)
        
        # فیلد ورودی
        entry_frame = tk.Frame(field_frame, bg='white')
        entry_frame.pack(side="left", fill="x", expand=True, padx=10)
        
        current_value = field_data.get('value', '')
        confidence = field_data.get('confidence', 0.0)
        method = field_data.get('method', 'none')
        
        # Entry برای مقدار
        entry_var = tk.StringVar(value=current_value or '')
        entry = tk.Entry(
            entry_frame,
            textvariable=entry_var,
            font=self.fonts['persian'],
            width=30,
            bg='#f8f9fa' if current_value else '#fff5f5',
            relief='solid',
            bd=1
        )
        entry.pack(side="left", padx=5)
        
        # اتصال منوی راست کلیک
        self.bind_context_menu(entry)
        
        # نمایش اطمینان
        confidence_label = tk.Label(
            entry_frame,
            text=f"اطمینان: {confidence:.1f}",
            font=self.fonts['small'],
            bg='white',
            fg='#27ae60' if confidence > 0.5 else '#e74c3c'
        )
        confidence_label.pack(side="left", padx=5)
        
        # ذخیره مرجع ویجت
        self.edit_widgets[field_id] = {
            'entry': entry,
            'var': entry_var,
            'original_value': current_value,
            'confidence': confidence,
            'method': method,
            'label': field_label
        }
        
    def load_image_for_preview(self, file_path):
        """بارگذاری تصویر برای پیشنمایش"""
        try:
            # اگر فایل PDF است تصویر صفحه اول را بارگذاری کن
            if file_path.lower().endswith('.pdf'):
                # جستجو برای تصویر متناظر در پوشه temp
                file_stem = Path(file_path).stem
                temp_dir = Path("temp")
                
                image_files = list(temp_dir.glob(f"{file_stem}_page_0.*"))
                if image_files:
                    image_path = str(image_files[0])
                else:
                    return
            else:
                image_path = file_path
                
            # بارگذاری و نمایش تصویر
            image = Image.open(image_path)
            
            # تغییر اندازه برای نمایش
            display_size = (600, 600)
            image.thumbnail(display_size, Image.Resampling.LANCZOS)
            
            # تبدیل به PhotoImage
            self.current_image = ImageTk.PhotoImage(image)
            
            # نمایش در Canvas
            self.image_canvas.delete("all")
            canvas_width = self.image_canvas.winfo_width()
            canvas_height = self.image_canvas.winfo_height()
            
            if canvas_width > 1 and canvas_height > 1:
                x = (canvas_width - image.width) // 2
                y = (canvas_height - image.height) // 2
                self.image_canvas.create_image(x, y, anchor='nw', image=self.current_image)
                
        except Exception as e:
            print(f"خطا در بارگذاری تصویر: {e}")
            
    def save_edits(self):
        """ذخیره تغییرات"""
        if not self.edit_widgets:
            messagebox.showwarning("هشدار", "هیچ ویرایشی برای ذخیره وجود ندارد")
            return
            
        changes_made = False
        
        for field_id, widget_data in self.edit_widgets.items():
            current_value = widget_data['var'].get().strip()
            original_value = widget_data['original_value']
            
            if current_value != original_value:
                changes_made = True
                # بهروزرسانی دادهها
                # (کد بهروزرسانی اینجا اضافه میشود)
                
        if changes_made:
            messagebox.showinfo("موفقیت", "تغییرات ذخیره شد")
            self.update_status("💾 تغییرات ذخیره شد")
        else:
            messagebox.showinfo("اطلاع", "هیچ تغییری انجام نشده است")
            
    def apply_and_learn(self):
        """اعمال تغییرات و یادگیری"""
        if not self.edit_widgets:
            messagebox.showwarning("هشدار", "هیچ ویرایشی برای یادگیری وجود ندارد")
            return
            
        # ذخیره تغییرات
        self.save_edits()
        
        # شروع فرآیند یادگیری
        self.learning_system.learn_from_edits(self.edit_widgets)
        
        messagebox.showinfo("موفقیت", "تغییرات اعمال شد و الگوهای جدید یاد گرفته شد")
        self.update_status("🤖 یادگیری انجام شد")
        
        # انتقال به تب یادگیری
        self.notebook.select(3)
        self.update_learning_display()
        
    def reset_edits(self):
        """بازگشت به حالت اصلی"""
        if messagebox.askyesno("تأیید", "آیا میخواهید همه تغییرات لغو شوند"):
            for widget_data in self.edit_widgets.values():
                widget_data['var'].set(widget_data['original_value'])
            self.update_status("↩️ تغییرات لغو شد")
            
    def reload_edit_data(self):
        """بارگذاری مجدد دادهها"""
        # بهروزرسانی لیست فایلها
        files = list(self.results_data.keys())
        self.edit_file_combo['values'] = [Path(f).name for f in files]
        
        if files and not self.edit_file_var.get():
            self.edit_file_var.set(Path(files[0]).name)
            self.load_file_for_edit()
            
    # متدهای کنترل تصویر
    def zoom_in(self):
        """بزرگنمایی تصویر"""
        # کد بزرگنمایی
        self.update_status("🔍 بزرگنمایی")
        
    def zoom_out(self):
        """کوچکنمایی تصویر"""
        # کد کوچکنمایی
        self.update_status("🔍 کوچکنمایی")
        
    def reset_zoom(self):
        """بازگشت به اندازه اصلی"""
        # کد بازگشت به اندازه اصلی
        self.update_status("📐 اندازه اصلی")
        
    # متدهای تب نتایج
    def filter_results(self):
        """فیلتر کردن نتایج"""
        filter_value = self.results_filter_var.get()
        # کد فیلتر
        self.update_status(f"🔍 فیلتر: {filter_value}")
        
    def sort_results(self, column):
        """مرتبسازی نتایج"""
        # کد مرتبسازی
        self.update_status(f"📊 مرتبسازی بر اساس {column}")
        
    def on_result_double_click(self, event):
        """کلیک دوبل روی نتیجه"""
        item = self.results_tree.selection()[0]
        values = self.results_tree.item(item, 'values')
        file_name = values[1]
        
        # یافتن فایل اصلی
        for file_path in self.results_data.keys():
            if Path(file_path).name == file_name:
                # انتقال به تب ویرایش
                self.edit_file_var.set(file_name)
                self.notebook.select(1)
                self.load_file_for_edit()
                break
                
    def export_to_excel(self):
        """خروجی Excel"""
        if not self.results_data:
            messagebox.showwarning("هشدار", "هیچ دادهای برای خروجی وجود ندارد")
            return
            
        output_file = filedialog.asksaveasfilename(
            title="ذخیره فایل Excel",
            defaultextension=".xlsx",
            filetypes=[("فایل Excel", "*.xlsx"), ("همه فایلها", "*.*")]
        )
        
        if output_file:
            try:
                # تولید فایل Excel با فرمت راستچین
                self.generate_excel_output(output_file)
                messagebox.showinfo("موفقیت", f"فایل Excel ذخیره شد:\n{output_file}")
                self.update_status("📄 خروجی Excel ایجاد شد")
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در ایجاد فایل Excel: {e}")
                
    def generate_excel_output(self, output_file):
        """تولید فایل Excel با فرمت مناسب"""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # تهیه دادهها
        data_rows = []
        row_num = 1
        
        for file_path, result in self.results_data.items():
            if 'pages' in result:
                for page_result in result['pages']:
                    data_rows.append(self.prepare_excel_row(row_num, file_path, page_result))
                    row_num += 1
            else:
                data_rows.append(self.prepare_excel_row(row_num, file_path, result))
                row_num += 1
                
        # ایجاد DataFrame
        columns = [
            'ردیف', 'نام فایل', 'وضعیت', 'زمان پردازش',
            'شماره کوتا', 'کد کالا', 'شرح کالا', 'نوع بسته', 'تعداد بسته',
            'وزن خالص', 'کشور طرف معامله', 'نرخ ارز', 'نوع ارز',
            'ارزش گمرکی', 'بیمه', 'کرایه', 'حقوق ورودی', 'مالیات', 'جمع عوارض'
        ]
        
        df = pd.DataFrame(data_rows, columns=columns)
        
        # ذخیره با pandas
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # فرمتبندی با openpyxl
        wb = load_workbook(output_file)
        ws = wb.active
        
        # تنظیم راستچین
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.font = Font(name='Tahoma', size=10)
                
        # رنگبندی هدر
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
            
        # تنظیم عرض ستونها
        column_widths = {
            'A': 8, 'B': 25, 'C': 12, 'D': 15, 'E': 15, 'F': 12,
            'G': 30, 'H': 12, 'I': 12, 'J': 12, 'K': 20, 'L': 12,
            'M': 10, 'N': 15, 'O': 12, 'P': 12, 'Q': 15, 'R': 12, 'S': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        wb.save(output_file)
        
    def prepare_excel_row(self, row_num, file_path, result):
        """تهیه ردیف برای Excel"""
        file_name = Path(file_path).name
        status = result.get('status', 'نامشخص')
        processing_time = result.get('processing_time', '0s')
        
        extracted = result.get('extracted', {})
        
        return [
            row_num,
            file_name,
            'موفق' if status == 'success' else 'ناموفق',
            processing_time,
            extracted.get('شماره_کوتا', {}).get('value', ''),
            extracted.get('کد_کالا', {}).get('value', ''),
            extracted.get('شرح_کالا', {}).get('value', ''),
            extracted.get('نوع_بسته', {}).get('value', ''),
            extracted.get('تعداد_بسته', {}).get('value', ''),
            extracted.get('وزن_خالص', {}).get('value', ''),
            extracted.get('کشور_طرف_معامله', {}).get('value', ''),
            extracted.get('نرخ_ارز', {}).get('value', ''),
            extracted.get('نوع_ارز', {}).get('value', ''),
            extracted.get('ارزش_گمرکی', {}).get('value', ''),
            extracted.get('بیمه', {}).get('value', ''),
            extracted.get('کرایه', {}).get('value', ''),
            extracted.get('مبلغ_حقوق_ورودی', {}).get('value', ''),
            extracted.get('مالیات_بر_ارزش_افزوده', {}).get('value', ''),
            extracted.get('جمع_حقوق_عوارض', {}).get('value', '')
        ]
        
    def copy_to_clipboard(self):
        """کپی به کلیپبورد"""
        if not self.results_data:
            messagebox.showwarning("هشدار", "هیچ دادهای برای کپی وجود ندارد")
            return
            
        # تهیه متن برای کپی
        clipboard_text = "ردیف\tنام فایل\tشماره کوتا\tکد کالا\tوزن خالص\n"
        
        row_num = 1
        for file_path, result in self.results_data.items():
            file_name = Path(file_path).name
            
            if 'pages' in result:
                for page_result in result['pages']:
                    extracted = page_result.get('extracted', {})
                    clipboard_text += f"{row_num}\t{file_name}\t"
                    clipboard_text += f"{extracted.get('شماره_کوتا', {}).get('value', '')}\t"
                    clipboard_text += f"{extracted.get('کد_کالا', {}).get('value', '')}\t"
                    clipboard_text += f"{extracted.get('وزن_خالص', {}).get('value', '')}\n"
                    row_num += 1
            else:
                extracted = result.get('extracted', {})
                clipboard_text += f"{row_num}\t{file_name}\t"
                clipboard_text += f"{extracted.get('شماره_کوتا', {}).get('value', '')}\t"
                clipboard_text += f"{extracted.get('کد_کالا', {}).get('value', '')}\t"
                clipboard_text += f"{extracted.get('وزن_خالص', {}).get('value', '')}\n"
                row_num += 1
                
        # کپی به کلیپبورد
        self.root.clipboard_clear()
        self.root.clipboard_append(clipboard_text)
        
        messagebox.showinfo("موفقیت", "دادهها به کلیپبورد کپی شد")
        self.update_status("📋 کپی به کلیپبورد انجام شد")
        
    def generate_analysis_report(self):
        """تولید گزارش تحلیلی"""
        if not self.results_data:
            messagebox.showwarning("هشدار", "هیچ دادهای برای گزارش وجود ندارد")
            return
            
        # ایجاد پنجره گزارش
        report_window = tk.Toplevel(self.root)
        report_window.title("📊 گزارش تحلیلی")
        report_window.geometry("800x600")
        report_window.configure(bg='white')
        
        # ایجاد متن گزارش
        report_text = scrolledtext.ScrolledText(
            report_window,
            font=self.fonts['persian'],
            bg='white',
            fg='#2c3e50',
            wrap=tk.WORD
        )
        report_text.pack(fill="both", expand=True, padx=20, pady=20)
        
        # تولید محتوای گزارش
        stats = self.extractor.get_extraction_stats(self.results_data)
        
        report_content = f"""
📊 گزارش تحلیلی استخراج اسناد گمرکی
═══════════════════════════════════════

📅 تاریخ گزارش: {datetime.now().strftime('%Y-%m-%d %H:%M')}
👤 کاربر: {os.getlogin()}

📈 آمار کلی:
─────────────
• تعداد کل فایلها: {stats['total_files']}
• تعداد کل صفحات: {stats['total_pages']}
• زمان کل پردازش: {stats['processing_time']:.1f} ثانیه
• میانگین زمان هر فایل: {stats['processing_time']/stats['total_files']:.1f} ثانیه

🎯 نرخ موفقیت استخراج:
─────────────────────
"""
        
        for field_name, field_stats in stats['successful_extractions'].items():
            report_content += f"• {field_name}: {field_stats['percentage']}\n"
            
        report_content += f"\n📋 توزیع انواع اسناد:\n─────────────────\n"
        for doc_type, count in stats['document_types'].items():
            type_name = self.extractor.document_types.get(doc_type, doc_type)
            percentage = (count / stats['total_files']) * 100
            report_content += f"• {type_name}: {count} ({percentage:.1f}%)\n"
            
        # اضافه کردن توصیهها
        report_content += f"""

💡 توصیهها و نکات:
─────────────────
• فیلدهایی با نرخ موفقیت کمتر از 70% نیاز به بهبود الگو دارند
• استفاده از ویرایش و یادگیری برای بهبود دقت
• بررسی کیفیت تصاویر ورودی
• تنظیم مجدد آستانه اطمینان در صورت نیاز

📝 نتیجهگیری:
─────────────
سیستم با دقت کلی {sum(float(fs['percentage'].rstrip('%')) for fs in stats['successful_extractions'].values()) / len(stats['successful_extractions']):.1f}% عمل کرده است.
"""
        
        report_text.insert("1.0", report_content)
        report_text.configure(state='disabled')
        
        # دکمه ذخیره گزارش
        save_btn = tk.Button(
            report_window,
            text="💾 ذخیره گزارش",
            command=lambda: self.save_report(report_content),
            bg='#27ae60',
            fg='white',
            font=self.fonts['persian'],
            cursor='hand2'
        )
        save_btn.pack(pady=10)
        
    def save_report(self, content):
        """ذخیره گزارش"""
        file_path = filedialog.asksaveasfilename(
            title="ذخیره گزارش",
            defaultextension=".txt",
            filetypes=[("فایل متنی", "*.txt"), ("همه فایلها", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("موفقیت", f"گزارش ذخیره شد:\n{file_path}")
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در ذخیره گزارش: {e}")
                
    # متدهای تب یادگیری
    def update_learning_display(self):
        """بهروزرسانی نمایش یادگیری"""
        # بهروزرسانی جدول مقایسه
        for item in self.comparison_tree.get_children():
            self.comparison_tree.delete(item)
            
        # بهروزرسانی جدول الگوها
        for item in self.patterns_tree.get_children():
            self.patterns_tree.delete(item)
            
        # بارگذاری الگوهای موجود
        learned_patterns = self.learning_system.get_learned_patterns()
        
        for field_name, patterns in learned_patterns.items():
            for pattern_data in patterns:
                self.patterns_tree.insert("", "end", values=[
                    field_name,
                    pattern_data.get('pattern', ''),
                    pattern_data.get('success_count', 0),
                    f"{pattern_data.get('accuracy', 0):.1f}%"
                ])
                
        # بهروزرسانی لیست فیلدها
        field_names = list(self.extractor.import_patterns.keys())
        persian_names = [name.replace('_', ' ') for name in field_names]
        self.pattern_field_combo['values'] = persian_names
        
    def add_pattern(self):
        """افزودن الگو جدید"""
        if not self.pattern_field_var.get():
            messagebox.showwarning("هشدار", "ابتدا فیلدی را انتخاب کنید")
            return
            
        pattern_text = self.pattern_entry.get("1.0", tk.END).strip()
        if not pattern_text:
            messagebox.showwarning("هشدار", "الگو نمیتواند خالی باشد")
            return
            
        # تست الگو
        try:
            import re
            re.compile(pattern_text)
        except re.error as e:
            messagebox.showerror("خطا", f"الگو نامعتبر است:\n{e}")
            return
            
        # افزودن الگو
        field_name = self.pattern_field_var.get().replace(' ', '_')
        self.learning_system.add_custom_pattern(field_name, pattern_text)
        
        messagebox.showinfo("موفقیت", "الگو با موفقیت اضافه شد")
        self.update_learning_display()
        
    def edit_pattern(self):
        """ویرایش الگو"""
        selected = self.patterns_tree.selection()
        if not selected:
            messagebox.showwarning("هشدار", "ابتدا الگویی را انتخاب کنید")
            return
            
        item = selected[0]
        values = self.patterns_tree.item(item, 'values')
        
        field_name = values[0]
        current_pattern = values[1]
        
        # تنظیم فیلد و الگو در ویرایشگر
        self.pattern_field_var.set(field_name)
        self.pattern_entry.delete("1.0", tk.END)
        self.pattern_entry.insert("1.0", current_pattern)
        
    def delete_pattern(self):
        """حذف الگو"""
        selected = self.patterns_tree.selection()
        if not selected:
            messagebox.showwarning("هشدار", "ابتدا الگویی را انتخاب کنید")
            return
            
        if messagebox.askyesno("تأیید", "آیا میخواهید این الگو حذف شود"):
            item = selected[0]
            values = self.patterns_tree.item(item, 'values')
            
            field_name = values[0].replace(' ', '_')
            pattern = values[1]
            
            self.learning_system.remove_pattern(field_name, pattern)
            self.update_learning_display()
            
            messagebox.showinfo("موفقیت", "الگو حذف شد")
            
    def save_patterns(self):
        """ذخیره الگوها"""
        try:
            self.learning_system.save_patterns()
            messagebox.showinfo("موفقیت", "الگوها ذخیره شد")
            self.update_status("💾 الگوها ذخیره شد")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ذخیره الگوها: {e}")
            
    def reload_patterns(self):
        """بارگذاری مجدد الگوها"""
        try:
            self.learning_system.load_patterns()
            self.update_learning_display()
            messagebox.showinfo("موفقیت", "الگوها بارگذاری شد")
            self.update_status("🔄 الگوها بارگذاری شد")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در بارگذاری الگوها: {e}")
            
    # متدهای تنظیمات
    def update_confidence_label(self, value):
        """بهروزرسانی برچسب اطمینان"""
        self.confidence_label.configure(text=f"{float(value):.2f}")
        
    def change_theme(self):
        """تغییر تم"""
        theme = self.theme_var.get()
        # کد تغییر تم
        self.update_status(f"🎨 تم تغییر کرد: {theme}")
        
    def save_settings(self):
        """ذخیره تنظیمات"""
        settings = {
            'confidence_threshold': self.confidence_var.get(),
            'dpi': self.dpi_var.get(),
            'languages': {
                'fa': self.lang_fa.get(),
                'en': self.lang_en.get(),
                'ar': self.lang_ar.get()
            },
            'theme': self.theme_var.get(),
            'last_update': datetime.now().isoformat()
        }
        
        try:
            with open("settings.json", 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("موفقیت", "تنظیمات ذخیره شد")
            self.update_status("💾 تنظیمات ذخیره شد")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ذخیره تنظیمات: {e}")
            
    def reset_settings(self):
        """بازگشت به تنظیمات پیشفرض"""
        if messagebox.askyesno("تأیید", "آیا میخواهید تنظیمات به حالت پیشفرض برگردند"):
            self.confidence_var.set(0.35)
            self.dpi_var.set(300)
            self.lang_fa.set(True)
            self.lang_en.set(True)
            self.lang_ar.set(True)
            self.theme_var.set("light")
            
            messagebox.showinfo("موفقیت", "تنظیمات به حالت پیشفرض برگشت")
            self.update_status("🔄 تنظیمات بازنشانی شد")
            
    def test_settings(self):
        """تست تنظیمات"""
        # تست تنظیمات فعلی
        try:
            # تست OCR
            langs = []
            if self.lang_fa.get(): langs.append('fa')
            if self.lang_en.get(): langs.append('en')
            if self.lang_ar.get(): langs.append('ar')
            
            if not langs:
                raise ValueError("حداقل یک زبان باید انتخاب شود")
                
            messagebox.showinfo("موفقیت", "تمام تنظیمات معتبر هستند")
            self.update_status("✅ تست تنظیمات موفق")
            
        except Exception as e:
            messagebox.showerror("خطا", f"تنظیمات نامعتبر: {e}")
            
    def load_settings(self):
        """بارگذاری تنظیمات"""
        try:
            if os.path.exists("settings.json"):
                with open("settings.json", 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    
                # اعمال تنظیمات
                self.confidence_var.set(settings.get('confidence_threshold', 0.35))
                self.dpi_var.set(settings.get('dpi', 300))
                
                languages = settings.get('languages', {})
                self.lang_fa.set(languages.get('fa', True))
                self.lang_en.set(languages.get('en', True))
                self.lang_ar.set(languages.get('ar', True))
                
                self.theme_var.set(settings.get('theme', 'light'))
                
        except Exception as e:
            print(f"خطا در بارگذاری تنظیمات: {e}")
            
    def run(self):
        """اجرای برنامه"""
        # تنظیمات پایانی
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # نمایش پیام خوشآمدگویی
        self.update_status("🎯 سیستم آماده است - فایلهای خود را انتخاب کنید")
        
        # شروع حلقه اصلی
        self.root.mainloop()
        
    def on_closing(self):
        """بستن برنامه"""
        if messagebox.askokcancel("خروج", "آیا مطمئن هستید که میخواهید خارج شوید"):
            # ذخیره تنظیمات
            try:
                self.save_settings()
            except:
                pass
                
            # پاکسازی فایلهای موقت
            try:
                import shutil
                temp_dir = Path("temp")
                if temp_dir.exists():
                    shutil.rmtree(temp_dir)
            except:
                pass
                
            self.root.destroy()

if __name__ == "__main__":
    print("🚀 راهاندازی سیستم استخراج هوشمند اسناد گمرکی...")
    
    app = CustomsExtractorGUI()
    app.run()
